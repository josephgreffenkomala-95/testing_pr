from __future__ import annotations

from decimal import Decimal
from typing import Iterable, Protocol

from finance_manager.models.entities import Account, Budget, Category, PlannedTransaction, Transaction
from finance_manager.models.schemas import SHEET_HEADERS, entity_to_row
from finance_manager.services.sheets import GoogleSheetsRepository


MONTH_NAME_TO_NUMBER = {
    "januari": 1, "jan": 1,
    "februari": 2, "feb": 2,
    "maret": 3, "mar": 3,
    "april": 4, "apr": 4,
    "mei": 5,
    "juni": 6, "jun": 6,
    "juli": 7, "jul": 7,
    "agustus": 8, "aug": 8, "ags": 8,
    "september": 9, "sep": 9, "sept": 9,
    "oktober": 10, "okt": 10, "oct": 10,
    "november": 11, "nov": 11,
    "desember": 12, "des": 12, "dec": 12,
}


class CellLike(Protocol):
    value: object


class CellGrid(Protocol):
    def cell(self, row: int, column: int) -> CellLike: ...


class MagangPlan:
    """Parsed view of the `magang` sheet inside moneyyy.xlsx.

    The sheet is loosely structured and uses Indonesian labels. This parser
    extracts the four blocks that represent the user's financial plan:

    - Gaji magang (internship income) at A6:D16
    - Pengeluaran perbulan (monthly recurring expenses) at J6:K15
    - Kantong bersama (shared-pocket expenses) at R6:S11
    - Perpuluhan Mami (monthly tithe to mother) at B23:D29
    """

    def __init__(self, grid: CellGrid) -> None:
        self._grid = grid
        self.salary_rows: list[SalaryRow] = self._read_salary()
        self.monthly_expenses: list[LineItem] = self._read_block(start_row=6, end_row=15, name_col=10, amount_col=11)
        self.shared_pocket: list[LineItem] = self._read_block(start_row=6, end_row=11, name_col=18, amount_col=19)
        self.tithe_rows: list[TitheRow] = self._read_tithe()

    def _cell(self, row: int, column: int) -> object:
        return self._grid.cell(row, column).value

    def _read_salary(self) -> list[SalaryRow]:
        rows: list[SalaryRow] = []
        for r in range(6, 17):
            index = self._cell(r, 1)
            month_label = self._cell(r, 2)
            amount = self._cell(r, 3)
            status = self._cell(r, 4)
            if index is None and month_label is None and amount is None:
                continue
            rows.append(SalaryRow(index, str(month_label or "").strip(), amount, str(status or "").strip().lower()))
        return rows

    def _read_block(self, start_row: int, end_row: int, name_col: int, amount_col: int) -> list[LineItem]:
        items: list[LineItem] = []
        for r in range(start_row, end_row + 1):
            name = self._cell(r, name_col)
            amount = self._cell(r, amount_col)
            if not name:
                continue
            items.append(LineItem(str(name).strip(), amount))
        return items

    def _read_tithe(self) -> list[TitheRow]:
        rows: list[TitheRow] = []
        for r in range(23, 30):
            index = self._cell(r, 2)
            amount = self._cell(r, 3)
            status = self._cell(r, 4)
            if index is None and amount is None:
                continue
            rows.append(TitheRow(index, amount, str(status or "").strip().lower()))
        return rows


class SalaryRow:
    def __init__(self, index: object, month_label: str, amount: object, status: str) -> None:
        self.index = index
        self.month_label = month_label
        self.amount = _to_decimal(amount)
        self.status = status

    @property
    def is_done(self) -> bool:
        return self.status == "done"

    @property
    def month_number(self) -> int | None:
        if not self.month_label:
            return None
        return MONTH_NAME_TO_NUMBER.get(self.month_label.lower())

    def description(self, year: int) -> str:
        return f"Gaji magang {self.month_label}".strip()


class TitheRow:
    def __init__(self, index: object, amount: object, status: str) -> None:
        self.index = index
        self.amount = _to_decimal(amount)
        self.status = status

    @property
    def is_done(self) -> bool:
        return self.status.startswith("done")


class LineItem:
    def __init__(self, name: str, amount: object) -> None:
        self.name = name
        self.amount = _to_decimal(amount)


def _to_decimal(value: object) -> Decimal:
    if value is None:
        return Decimal("0")
    try:
        return Decimal(str(value)).quantize(Decimal("0.01"))
    except Exception:
        return Decimal("0")


def _default_year() -> int:
    from datetime import UTC, datetime
    return datetime.now(UTC).year


def migrate_magang(
    repository: GoogleSheetsRepository,
    plan: MagangPlan,
    *,
    year: int | None = None,
    account_name: str = "Cash",
    tithe_account_name: str = "Cash",
) -> MigrationReport:
    year = year or _default_year()
    report = MigrationReport()
    state = _MigrationState(repository)

    cash_account = state.ensure_account(account_name)
    tithe_account = state.ensure_account(tithe_account_name)

    for item in plan.monthly_expenses:
        if item.amount <= 0:
            continue
        budget = state.append_budget(item.name, item.amount, year)
        if budget is not None:
            report.budgets_created += 1
    for item in plan.shared_pocket:
        if item.amount <= 0:
            continue
        budget = state.append_budget(f"{item.name} (Kantong Bersama)", item.amount, year)
        if budget is not None:
            report.budgets_created += 1

    for row in plan.salary_rows:
        if row.amount <= 0:
            continue
        month_number = row.month_number
        if month_number is None:
            continue
        category = state.ensure_category("Salary", "income")
        if row.is_done:
            transaction = state.append_transaction(
                entry_type="income",
                date=_date(year, month_number, 1),
                amount=row.amount,
                category=category,
                account=cash_account,
                description=row.description(year),
            )
            if transaction is not None:
                report.transactions_created += 1
                report.transactions.append(transaction)
        else:
            planned = state.append_planned(
                entry_type="income",
                status="planned",
                expected_date=_date(year, month_number, 1),
                amount=row.amount,
                category=category,
                account=cash_account,
                description=row.description(year),
            )
            if planned is not None:
                report.planned_created += 1
                report.planned.append(planned)

    tithe_category = state.ensure_category("Perpuluhan Mami", "expense")
    for index, tithe in enumerate(plan.tithe_rows, start=1):
        if tithe.amount <= 0:
            continue
        month_number = _tithe_month_number(index)
        if tithe.is_done:
            transaction = state.append_transaction(
                entry_type="expense",
                date=_date(year, month_number, 1),
                amount=tithe.amount,
                category=tithe_category,
                account=tithe_account,
                description=f"Perpuluhan Mami ke-{index}",
            )
            if transaction is not None:
                report.transactions_created += 1
                report.transactions.append(transaction)
        else:
            planned = state.append_planned(
                entry_type="expense",
                status="planned",
                expected_date=_date(year, month_number, 1),
                amount=tithe.amount,
                category=tithe_category,
                account=tithe_account,
                description=f"Perpuluhan Mami ke-{index}",
            )
            if planned is not None:
                report.planned_created += 1
                report.planned.append(planned)

    return report


def _tithe_month_number(index: int) -> int:
    base = 2
    return base + (index - 1)


def _date(year: int, month: int, day: int) -> str:
    return f"{year:04d}-{month:02d}-{day:02d}"


class _MigrationState:
    def __init__(self, repository: GoogleSheetsRepository) -> None:
        from datetime import UTC, datetime

        self.repository = repository
        self.snapshot = repository.load_snapshot()
        self.now = repository._now()
        self.current_month = datetime.now(UTC).strftime("%Y-%m")
        self.worksheets = {worksheet.title: worksheet for worksheet in repository._spreadsheet.worksheets()}
        self.accounts_by_name = {
            account.name.strip().lower(): account
            for account in self.snapshot.accounts
            if account.name.strip()
        }
        self.categories_by_key = {
            (category.name.strip().lower(), category.kind): category
            for category in self.snapshot.categories
            if category.name.strip()
        }
        self.transaction_keys = {
            (item.entry_type, item.date, item.amount, item.category_id, item.account_id, item.description)
            for item in self.snapshot.transactions
        }
        self.planned_keys = {
            (item.entry_type, item.status, item.expected_date or "", item.amount, item.category_id, item.account_id, item.description)
            for item in self.snapshot.planned_transactions
        }
        self.budget_keys = {
            (item.month, item.category_id, item.entry_type)
            for item in self.snapshot.budgets
        }
        self.account_index = _highest_suffix(self.snapshot.accounts, "ACC")
        self.category_index = _highest_suffix(self.snapshot.categories, "CAT")
        self.transaction_index = _highest_suffix(self.snapshot.transactions, "TX")
        self.planned_index = _highest_suffix(self.snapshot.planned_transactions, "PLN")
        self.budget_index = _highest_suffix(self.snapshot.budgets, "BDG")

    def ensure_account(self, name: str) -> Account:
        key = name.strip().lower()
        existing = self.accounts_by_name.get(key)
        if existing is not None:
            return existing
        self.account_index += 1
        account = Account(
            id=f"ACC{self.account_index:04d}",
            name=name.strip(),
            account_type="cash",
            currency="IDR",
            current_balance=Decimal("0.00"),
            is_active=True,
        )
        self._append_entity("Accounts", account)
        self.accounts_by_name[key] = account
        return account

    def ensure_category(self, name: str, entry_type: str) -> Category:
        key = (name.strip().lower(), entry_type)
        existing = self.categories_by_key.get(key)
        if existing is not None:
            return existing
        self.category_index += 1
        category = Category(
            id=f"CAT{self.category_index:04d}",
            name=name.strip(),
            kind=entry_type,
            is_active=True,
        )
        self._append_entity("Categories", category)
        self.categories_by_key[key] = category
        return category

    def append_budget(self, category_name: str, amount: Decimal, year: int) -> Budget | None:
        category = self.ensure_category(category_name, "expense")
        key = (self.current_month, category.id, "expense")
        if key in self.budget_keys:
            return None
        self.budget_index += 1
        budget = Budget(
            id=f"BDG{self.budget_index:04d}",
            month=self.current_month,
            category_id=category.id,
            entry_type="expense",
            amount=amount,
            notes=f"Migrated from magang plan ({year})",
            created_at=self.now,
            updated_at=self.now,
        )
        self._append_entity("Budgets", budget)
        self.budget_keys.add(key)
        return budget

    def append_transaction(
        self,
        *,
        entry_type: str,
        date: str,
        amount: Decimal,
        category: Category,
        account: Account,
        description: str,
    ) -> Transaction | None:
        key = (entry_type, date, amount, category.id, account.id, description)
        if key in self.transaction_keys:
            return None
        self.transaction_index += 1
        transaction = Transaction(
            id=f"TX{self.transaction_index:04d}",
            entry_type=entry_type,
            date=date,
            amount=amount,
            category_id=category.id,
            account_id=account.id,
            description=description,
            notes="Migrated from magang sheet",
            created_at=self.now,
            updated_at=self.now,
        )
        self._append_entity("Transactions", transaction)
        self.transaction_keys.add(key)
        return transaction

    def append_planned(
        self,
        *,
        entry_type: str,
        status: str,
        expected_date: str,
        amount: Decimal,
        category: Category,
        account: Account,
        description: str,
    ) -> PlannedTransaction | None:
        key = (entry_type, status, expected_date, amount, category.id, account.id, description)
        if key in self.planned_keys:
            return None
        self.planned_index += 1
        planned = PlannedTransaction(
            id=f"PLN{self.planned_index:04d}",
            entry_type=entry_type,
            status=status,
            expected_date=expected_date,
            amount=amount,
            category_id=category.id,
            account_id=account.id,
            description=description,
            notes="Migrated from magang sheet",
            created_at=self.now,
            updated_at=self.now,
        )
        self._append_entity("Planned Transactions", planned)
        self.planned_keys.add(key)
        return planned

    def _append_entity(self, title: str, entity: object) -> None:
        worksheet = self.worksheets[title]
        worksheet.append_row(entity_to_row(entity, SHEET_HEADERS[title]))


def _highest_suffix(entities: list[object], prefix: str) -> int:
    highest = 0
    for entity in entities:
        entity_id = getattr(entity, "id", "")
        if not isinstance(entity_id, str) or not entity_id.startswith(prefix):
            continue
        try:
            highest = max(highest, int(entity_id[len(prefix):]))
        except ValueError:
            continue
    return highest


class MigrationReport:
    def __init__(self) -> None:
        self.transactions_created = 0
        self.planned_created = 0
        self.budgets_created = 0
        self.transactions: list[Transaction] = []
        self.planned: list[PlannedTransaction] = []

    def summary(self) -> str:
        return (
            f"Migration complete: "
            f"{self.transactions_created} transactions, "
            f"{self.planned_created} planned transactions, "
            f"{self.budgets_created} budgets."
        )


def load_magang_from_path(path: str, sheet: str = "magang") -> MagangPlan:
    """Read the `magang` sheet from an .xlsx workbook on disk."""
    import openpyxl

    workbook = openpyxl.load_workbook(path, data_only=True)
    if sheet not in workbook.sheetnames:
        raise KeyError(f"Sheet '{sheet}' not found in {path}. Available: {workbook.sheetnames}")
    worksheet = workbook[sheet]
    return MagangPlan(worksheet)


def iter_magang_salary(rows: Iterable[SalaryRow]) -> Iterable[SalaryRow]:
    for row in rows:
        if row.amount and row.month_number:
            yield row
